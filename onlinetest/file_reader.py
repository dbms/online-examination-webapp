import csv
import openpyxl
import os
from .models import studentProfile, question, studentMark

root_path = os.path.dirname(os.path.abspath(__file__))

def file_to_db(filename, client_name, test_id):
    '''main function: write file to database'''
    print(filename[-3:] + 'sdaasd')
    if filename[-3:] == 'csv':
        data = read_csv(filename)
        write_db(data, client_name, test_id)        
    elif filename[-4:] == 'xlsx' or filename[-3:] == 'xls':
        data = read_xl(filename)
        write_db(data, client_name, test_id)
    else:
        print('error')
    
def read_csv(filename):
    '''read csv file'''
    with open(root_path + '/static/onlinetest/docs/' + filename) as csvfile:
        spamreader = csv.reader(csvfile, delimiter=' ', quotechar='|')
        data = {}
        data_row = []
        i = 0
        for row in spamreader:
            data[i] = {}
            j = 0
            data_row = row[0].split(',')
            for row_cell in data_row:
                a = row_cell 
                data[i][j] = a if a != None else 'na'
                j += 1
            i += 1
    data['filename'] = filename
        #print(data)
    return data

    
def read_xl(filename):
    '''read excel file'''
    wb = openpyxl.load_workbook(root_path + '/static/onlinetest/docs/' + filename)
    print(wb.get_sheet_names())
    anotherSheet = wb.active
    ws = wb.worksheets[0]
    #print(ws.read())
    #total sheets
    data = {}
    for i in wb.worksheets:
        #all rows and columns
        row_count = i.max_row +1 
        col_count = i.max_column
        print(row_count, col_count)
        for row in range(2,row_count):
            data[row-1] = {}
            for col in range(1,col_count):
                a = i.cell(column=col, row=row).value
                data[row-1][col-1] = a if a != None else 'na'
    data['filename'] = filename
    #print(data)
    return data

def write_db(data, client_name, file_id):
    '''write to database'''
    #ques_paper=quesFile.objects.get(ques_paper_id=filename, client=client_name)
    for i in data:
        if str(i) == 'filename':
            break
        ques=question.objects.create(
            question_id=file_id,
            question=data[i].get(1),
            option1=data[i].get(2),
            option2=data[i].get(3),
            option3=data[i].get(4),
            option4=data[i].get(5),
            answer=data[i].get(6),
            #questionType=data[i].get(7),
        )
    '''
    for i in data:
        if str(i) == 'filename':
            break
        print(data['filename'][:data['filename'].rfind('.')] + str(data[i].get(0)),
            ' ' + str(data[i].get(1)),
        data[i].get(2),
        data[i].get(3),
        data[i].get(4),
        data[i].get(5),
        data[i].get(6),
        data[i].get(7),
        )
    '''    
#read_xl('quesformat.xlsx')

#file_to_db('quesformat1.xlsx','yoyo')

